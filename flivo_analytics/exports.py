"""
Flivo Backlink Opportunity Analytics — Exports
=================================================
Writes the cleaned dataset, opportunity ranking, and all summary tables
into a fresh "Analysis Summary" workbook, and also appends key tables as
extra sheets onto a copy of the original research workbook.
"""
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from config import SOURCE_WORKBOOK, OUTPUT_DIR

NAVY = "1F3864"
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor=NAVY)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, index: bool = False, title: str = None) -> int:
    row = start_row
    if title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color=NAVY)
        row += 2

    df_to_write = df.reset_index() if index else df
    headers = list(df_to_write.columns)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=str(h))
        cell.font = header_font
        cell.fill = header_fill
    row += 1
    for _, record in df_to_write.iterrows():
        for c, val in enumerate(record, 1):
            if isinstance(val, float) and pd.isna(val):
                val = "N/A"
            ws.cell(row=row, column=c, value=val)
        row += 1

    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22
    return row + 2


def export_analysis_workbook(
    master_df, scored_df, pareto_df, kpi_df, da_stats, dr_stats, traffic_stats,
    pricing_df, guest_post_df, food_df, quick_wins_df, long_term_df, budget_df,
    correlation_df, chart_paths: list, output_filename: str = "Flivo_Backlink_Analysis_Report.xlsx"
):
    """Build a standalone analytics workbook: cleaned dataset, opportunity
    ranking, every summary table, and embedded charts."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    from openpyxl import Workbook
    wb = Workbook()

    # --- Sheet: KPI Summary ---
    ws = wb.active
    ws.title = "KPI Summary"
    _write_df(ws, kpi_df, title="SEO Opportunity — KPI Summary")

    # --- Sheet: Cleaned Master Dataset ---
    ws = wb.create_sheet("Cleaned Master Dataset")
    export_cols = ["Name", "Sheet", "Category", "Free_Paid", "DA_clean", "DR_clean",
                   "Traffic_clean", "Price_clean", "Is_Custom_Quote", "Guest_Post_Score",
                   "Food_Relevance_Score"]
    if "Cluster_Label" in master_df.columns:
        export_cols.append("Cluster_Label")
    _write_df(ws, master_df[export_cols], title="Cleaned & Standardized Dataset (all sheets combined)")

    # --- Sheet: Opportunity Ranking ---
    ws = wb.create_sheet("Opportunity Ranking")
    rank_cols = ["Rank", "Name", "Sheet", "Opportunity_Score", "Cumulative_Pct", "Partial_Data"]
    _write_df(ws, pareto_df[rank_cols], title="Weighted Opportunity Score Ranking (Pareto view)")

    # --- Sheet: Quick Wins & Long-Term ---
    ws = wb.create_sheet("Quick Wins & Long-Term")
    next_row = _write_df(ws, quick_wins_df, title="QUICK WINS — Free, high-score, guest-post-ready (act this month)")
    _write_df(ws, long_term_df, start_row=next_row, title="LONG-TERM OPPORTUNITIES — Paid, high-score (plan & budget)")

    # --- Sheet: Budget Plan ---
    ws = wb.create_sheet("Budget Optimization")
    note = f"Suggested allocation of a ${budget_df.attrs.get('budget', 'N/A')} monthly budget " \
           f"(spent: ${budget_df.attrs.get('spent', 'N/A')}, remaining: ${budget_df.attrs.get('remaining', 'N/A')})"
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=10)
    _write_df(ws, budget_df, start_row=3)

    # --- Sheet: Statistical Summaries ---
    ws = wb.create_sheet("DA-DR-Traffic Stats")
    row = 1
    ws.cell(row=row, column=1, value="Domain Authority Analysis").font = Font(bold=True, size=12, color=NAVY)
    row += 1
    for k, v in da_stats.items():
        if k == "top_5":
            continue
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        row += 1
    row += 2
    ws.cell(row=row, column=1, value="Domain Rating Analysis").font = Font(bold=True, size=12, color=NAVY)
    row += 1
    for k, v in dr_stats.items():
        if k == "top_5":
            continue
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        row += 1
    row += 2
    ws.cell(row=row, column=1, value="Traffic Analysis").font = Font(bold=True, size=12, color=NAVY)
    row += 1
    for k, v in traffic_stats.items():
        if k == "top_5":
            continue
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=str(v))
        row += 1
    row += 2
    row = _write_df(ws, pricing_df.reset_index(), start_row=row, title="Pricing Analysis by Category")
    row = _write_df(ws, guest_post_df.reset_index(), start_row=row, title="Guest Post Acceptance by Category")
    row = _write_df(ws, food_df.reset_index(), start_row=row, title="Food/Health Niche Suitability")
    _write_df(ws, correlation_df.reset_index(), start_row=row, title="Correlation Matrix (numeric fields)")

    # --- Sheet: Charts ---
    ws = wb.create_sheet("Charts")
    row_cursor = 1
    for path in chart_paths:
        if path and os.path.exists(path):
            img = XLImage(path)
            img.width, img.height = 520, 340
            ws.add_image(img, f"A{row_cursor}")
            row_cursor += 18

    output_path = os.path.join(OUTPUT_DIR, output_filename)
    wb.save(output_path)
    return output_path


def append_analysis_summary_to_original(kpi_df, pareto_df, output_filename: str = "Flivo_Backlink_Opportunity_Research.xlsx"):
    """Append a concise 'Analysis Summary' tab onto a fresh copy of the
    original research workbook, so both live in one place if preferred."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    dest = os.path.join(OUTPUT_DIR, output_filename)
    if not os.path.exists(dest):
        shutil.copy(SOURCE_WORKBOOK, dest)

    wb = load_workbook(dest)
    if "Analysis Summary" in wb.sheetnames:
        del wb["Analysis Summary"]
    ws = wb.create_sheet("Analysis Summary")
    _write_df(ws, kpi_df, title="SEO KPI Summary")
    top15 = pareto_df.head(15)[["Rank", "Name", "Sheet", "Opportunity_Score", "Partial_Data"]]
    _write_df(ws, top15, start_row=len(kpi_df) + 5, title="Top 15 by Weighted Opportunity Score (verified-data only)")
    wb.save(dest)
    return dest
